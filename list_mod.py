"""
Тут скоро будет описание скрипта.
Возможно даже на английском.

"""
import json
import codecs
from datetime import datetime, timedelta
from click import clear, secho
from openpyxl import load_workbook
from os import remove

# TODO: Добавь документацию. Missing function or method docstring (missing-function-docstring)
# TODO: цвет добавлен, отчиска добавленна, может что-то еще есть полезного в click?
#   https://click.palletsprojects.com/en/8.0.x/utils/
#   https://habr.com/ru/company/oleg-bunin/blog/551424/
# TODO: попробовать предусмотреть режим работы без цветов, а надо ли оно?
# TODO: довавить эксепшен если фаил уже открыт
#   предусмотреть в таком случае отключение меню редактирования

current_date = datetime.now().strftime('%d.%m.%Y')
cell_name = ['\nТемпература: ', 'Влажность: ', 'Давление: ', 'Напряжение: ', 'Частота: ']
weather_measure = [' °С', ' %', ' кПа', ' В', ' Гц']
format_list = ['н/д ', 'н/д ', 'н/д   ', 'н/д', 'н/д']


def check_and_input_config():
    try:
        with codecs.open('list_cfg.json', 'r', 'utf-8-sig') as f:
            config_file = json.load(f)
        return config_file
    except FileNotFoundError:
        config = {}
        secho('Конфиг list_cfg.json не найден, давай его создадим', bg='red', bold=True)
        config["file_address"] = input('\nВставте адрес папки (Ctrl + V): ') + '\\'
        config["file_name"] = input('Вставте имя файла (Ctrl + V): ')
        if '.xlsx' not in config["file_name"].lower():
            config["file_name"] += '.xlsx'
        print('Введи начальную строку поиска')
        start_search = input('По умолчанию значение 1000: ')
        if start_search != '':
            config["start_search"] = int(start_search)
        else:
            config["start_search"] = 1000
        print('Введи конечную строку поиска')
        end_search = input('По умолчанию значение 2001: ')
        if end_search != '':
            config["end_search"] = int(end_search)
        else:
            config["end_search"] = 2001
        with codecs.open('list_cfg.json', 'w+') as file:
            json.dump(config, file)
            print('Конфиг создан!')
    return config


def check_open_weather_file():
    try:
        open(config_params['file_address'] + config_params['file_name'], 'r')
        return True
    except PermissionError:
        secho('\nФайл с погодой уже кемто открыт!', bg='red', bold=True)
        input('Нажмите Enter для выхода ')
        return False
    except FileNotFoundError:
        secho('\nФайл с погодой не найден!', bg='red', bold=True)
        create_config = input('Пересоздать конфиг? y/n ')
        if create_config != 'y':
            return False
        else:
            remove('list_cfg.json')
            check_and_input_config()
            return True


def search_cell_address(search_date):
    value = ''
    for i in range(config_params["start_search"], config_params["end_search"], ):
        if search_date == str(ws['A' + str(i)].value.strftime('%d.%m.%Y')):
            value = str(i)
            break
    weather_cell_address = ['B' + value, 'C' + value, 'D' + value, 'E' + value, 'F' + value]
    return weather_cell_address


def print_weather(weather_cell_address, date=''):
    value = []
    for i in range(0, len(cell_name)):
        cell_value = str(ws[weather_cell_address[i]].value)
        if cell_value == 'None':
            cell_value = format_list[i]
        value.append(cell_name[i] + cell_value + weather_measure[i])
    if date != '':
        value[0] = value[0][1:]
    print(date, value[0], '|', value[1], '|', value[2], '|', value[3], '|', value[4])


def weather_today():
    clear()
    print(f'Норманьные условия сегодня: {current_date}')
    weather_cell_address = search_cell_address(current_date)
    print_weather(weather_cell_address)


def weather_input():
    clear()
    print('Ввод данных:')
    weather_cell_address = search_cell_address(current_date)
    for i in range(0, len(cell_name)):
        ws[weather_cell_address[i]] = float(input(cell_name[i]))
        ws[weather_cell_address[i]].number_format = '0.00'
    clear()
    print('Сохранение...')
    wb.save(config_params['file_address'] + config_params['file_name'])
    print(f'Сохранение выполнено!\n\nВведенные данные: {current_date}')
    print_weather(weather_cell_address)


def weather_by_date():
    clear()
    search_date = input('Введите дату в формате дд.мм.гггг: ')
    clear()
    print(f'Нормальные условия: {search_date}')
    weather_cell_address = search_cell_address(search_date)
    print_weather(weather_cell_address)


def weather_few_days():
    while True:
        try:
            day = int(input('За колько дней показать погоду? '))
            for i in range(0, day):
                back_date = (datetime.now() - timedelta(days=i)).strftime('%d.%m.%Y')
                weather_cell_address = search_cell_address(back_date)
                print('_' * 113)
                print_weather(weather_cell_address, back_date)
        except ValueError:
            clear()
            secho('Ошибка: всэ погано, ты ввел не число!', bg='red', bold=True)
            continue
        else:
            break


def waaagh():
    secho(r"""
                                              .__     ._.
    __  _  _______   _____   _____      ____  |  |__  | |
    \ \/ \/ /\__  \  \__  \  \__  \    / ___\ |  |  \ | |
     \     /  / __ \_ / __ \_ / __ \_ / /_/  >|   Y  \ \|
      \/\_/  (____  /(____  /(____  / \___  / |___|  / __
                  \/      \/      \/ /_____/       \/  \/
    """, fg='red')


clear()
secho(r"""
__        __              _    _                   ____       ___
\ \      / /  ___   __ _ | |_ | |__    ___  _ __  |___ \     / _ \
 \ \ /\ / /  / _ \ / _` || __|| '_ \  / _ \| '__|   __) |   | | | |
  \ V  V /  |  __/| (_| || |_ | | | ||  __/| |     / __/  _ | |_| |
   \_/\_/    \___| \__,_| \__||_| |_| \___||_|    |_____|(_) \___/
""", fg='yellow')

config_params = check_and_input_config()

try:
    open(config_params['file_address'] + config_params['file_name'], 'r')
    start_program = True
    
except PermissionError:
    secho('\nФайл с погодой уже кемто открыт!', bg='red', bold=True)
    input('\nНажмите Enter для выхода ')
    start_program = False
    
except FileNotFoundError:
    secho('\nФайл с погодой не найден!', bg='red', bold=True)
    create_or_die = input('\nПересоздать конфиг? y/n ')
    if create_or_die != 'y':
        start_program = False
        
    else:
        remove('list_cfg.json')
        config_params = check_and_input_config()
        start_program = True
            

while start_program:
    wb = load_workbook(config_params['file_address'] + config_params['file_name'])
    ws = wb.active
    secho('\n     Главное меню     ', bg='yellow', fg='black')
    print('\nДанные сегодня     - 0'
          '\nВвести данные      - 1'
          '\nДанные по дате     - 2'
          '\nДанные по дням     - 3'
          '\nПересоздать конфиг - +'
          '\nВыход              - 4')
    ans = input('\nВыберете действие: ')
    if ans == '4':
        break
    elif ans == '0':
        weather_today()
    elif ans == '1':
        weather_input()
    elif ans == '2':
        weather_by_date()
    elif ans == '3':
        clear()
        weather_few_days()
    elif ans == '+':
        clear()
        caution = input('Внимание! Это действие удалит текущий конфиг! Продолжить? y/n ')
        if caution == 'y':
            remove('list_cfg.json')
            config_params = check_and_input_config()
        else:
            continue
    elif ans == '7':
        clear()
        secho('НЕВЕРЫЙ ВВОД! Тут ничего нет, перестать тыкать не те кнопки!', bg='red', bold=True)
    elif ans == '9':
        clear()
        secho('Широкую на широкую!', bg='red', bold=True)
        waaagh()
    else:
        clear()
        secho('Ошибка: не правильный ввод', bg='red', bold=True)
