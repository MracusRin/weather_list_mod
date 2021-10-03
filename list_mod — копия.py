import openpyxl 
import time
from datetime import datetime

current_date = datetime.now().strftime('%d.%m.%Y')
#fileadd='Y:\\Супер общий зал\\Нормальные условия.xlsx'
fileadd='E:\\OneDrive\\Programming\\Python\\project\\exel\\weather.xlsx'
datacol=7 # Номер колонки где сохранена дата в файле
chort=0
start_search = 1000
end_searhc = 2001
wb = openpyxl.load_workbook(fileadd)
ws = wb.active    
title_parameter = ['Температура: ', 'Влажность: ', 'Давление: ', 'Напряжение: ', 'Частота: ']
#weather_unit = ['°С','%','кПа','В','Гц']
#Можно принимать значинея от пользователя в массивы!!!!!!!!

def search_line(search_date):
    #for i in range(1, ws.max_row + 1, ): # поиск по максимуму, до конца документа
    for i in range(start_search, end_searhc, ): #диапазаон поиска даты
        if search_date == ws.cell(i,datacol).value:
            data_line = str(ws.cell(i,datacol).row)
 #   tempcell = 
    humcell  = 'C' + data_line
    prescell = 'D' + data_line
    voltcell = 'E' + data_line
    freqcell = 'F' + data_line
    weather= ['B' + data_line ,humcell,prescell,voltcell,freqcell]
    return weather     

weather_cell=search_line(current_date) 
print(f'Норманьные условия сегодня: {current_date}')
print(f'\nТемпература: {ws[weather_cell[0]].value} °С | Влажность: {ws[weather_cell[1]].value} % | Давление: {ws[weather_cell[2]].value} кПа | Напряжение: {ws[weather_cell[3]].value} В | Частота: {ws[weather_cell[4]].value} Гц\n')

ans = input('Ввести данные - нажмите 1. Посмотреть данные по дате - нажмите 2: ')
if ans == '1':
    # можно вынести в начало и проверять условие там
    try: 
        myfile = open(fileadd, "r+") # or "a+", whatever you need
    except IOError:
        print ('\n!!! Какойто ЧОРТ уже открыл твой файл !!!\nВвод отмен')
        chort=1

    if chort != 1:

        for i in range(0, len(title_parameter)):
            ws[weather_cell[i]] = float(input(title_parameter[i]))
            ws[weather_cell[i]].number_format='0.0'           

        print('\nСохранение...')
        wb.save(fileadd)
        print('Сохранение выполнено!')

        print(f'\nВведенные данные: {current_date}')
        print(f'\nТемпература: {ws[weather_cell[0]].value} °С | Влажность: {ws[weather_cell[1]].value} % | Давление: {ws[weather_cell[2]].value} кПа | Напряжение: {ws[weather_cell[3]].value} В | Частота: {ws[weather_cell[4]].value} Гц\n')
elif ans == '2':
    search_date=input('Введите дату в формате дд.мм.гггг: ')
    weather_cell_seacrh = search_line(search_date) 
    print(f'\nНормальные условия: {search_date}')
    print(f'\nТемпература: {ws[weather_cell_seacrh[0]].value} °С | Влажность: {ws[weather_cell_seacrh[1]].value} % | Давление: {ws[weather_cell_seacrh[2]].value} кПа | Напряжение: {ws[weather_cell_seacrh[3]].value} В | Частота: {ws[weather_cell_seacrh[4]].value} Гц\n')
    time.sleep(7)
else:
    print('Ввод отмен')

time.sleep(3)

